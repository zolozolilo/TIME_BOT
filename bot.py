import logging
from datetime import datetime
from pathlib import Path
from queue import Queue
from typing import List

from telegram import Update
from telegram.ext import CommandHandler, ApplicationBuilder, ContextTypes

from openpyxl import Workbook, load_workbook

# Логирование
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Токен, полученный от BotFather
TOKEN = 'ВАШ ТОКЕН'

# Имя файла Excel
EXCEL_FILE = "logs.xlsx"

# Функция для инициализации Excel-файла
def initialize_excel_file() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(['User ID', 'Timestamp', 'Action'])
    workbook.save(EXCEL_FILE)

# Функция для добавления строки в Excel-файл
def append_to_excel(user_id: int, timestamp: str, action: str) -> None:
    excel_path = Path(EXCEL_FILE)
    if not excel_path.is_file():
        initialize_excel_file()

    try:
        workbook = load_workbook(excel_path)
        worksheet = workbook.active
        worksheet.append([user_id, timestamp, action])
        workbook.save(excel_path)
    except Exception as e:
        logger.error(f'Ошибка при записи в Excel: {e}')

# Функция для чтения данных из Excel-файла
def read_from_excel(user_id: int) -> List[List[str]]:
    excel_path = Path(EXCEL_FILE)
    if not excel_path.is_file():
        return []

    try:
        workbook = load_workbook(excel_path)
        worksheet = workbook.active
        rows = list(worksheet.rows)[1:]  # Пропустить заголовок
        filtered_rows = [[cell.value for cell in row] for row in rows if row[0].value == user_id]
        return filtered_rows
    except Exception as e:
        logger.error(f'Ошибка при чтении из Excel: {e}')
        return []

# Функция расчета общего рабочего времени
def calculate_work_time(user_id: int) -> str:
    rows = read_from_excel(user_id)
    total_seconds = 0

    for i in range(0, len(rows), 2):
        try:
            in_timestamp = datetime.strptime(rows[i][1], '%Y-%m-%d %H:%M:%S')
            out_timestamp = datetime.strptime(rows[i + 1][1], '%Y-%m-%d %H:%M:%S')
            delta = out_timestamp - in_timestamp
            total_seconds += delta.total_seconds()
        except IndexError:
            logger.warning(f'Для пользователя {user_id} нет записи ухода после прихода {rows[i][1]}')
            break

    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    return f"{hours} часов {minutes} минут"

# Команда старта
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text("Привет! Я бот для учета рабочего времени.")

# Команда помощи
async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text("Доступные команды:\n"
                                    "/in - отметить приход\n"
                                    "/out - отметить уход\n"
                                    "/summary - получить отчет по отработанным часам\n"
                                    "Введите команду, чтобы начать работу с ботом.")

# Команда регистрации прихода
async def check_in(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    append_to_excel(user_id, current_time, 'in')
    await update.message.reply_text(f"{update.effective_user.full_name}, вы отметили приход!")

# Команда регистрации ухода
async def check_out(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    append_to_excel(user_id, current_time, 'out')
    await update.message.reply_text("Вы отметили уход!")

# Команда получения отчета по отработанному времени
async def summary(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    work_time = calculate_work_time(user_id)
    await update.message.reply_text(f"Общее рабочее время: {work_time}")

# Основной блок программы
def main() -> None:
    application = ApplicationBuilder().token(TOKEN).build()

    application.add_handler(CommandHandler('start', start))
    application.add_handler(CommandHandler('help', help_command))
    application.add_handler(CommandHandler('in', check_in))
    application.add_handler(CommandHandler('out', check_out))
    application.add_handler(CommandHandler('summary', summary))

    application.run_polling()

if __name__ == '__main__':
    main()